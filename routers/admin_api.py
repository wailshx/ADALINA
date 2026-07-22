import os
import json
import secrets
import hashlib
import hmac as _hmac
import re
import time
import fcntl
import http.cookies
import subprocess
from typing import Optional, List, Any

from fastapi import APIRouter, Request, HTTPException, Depends, Body, Query, File, UploadFile, Form
from starlette.responses import JSONResponse as _StarletteJSONResponse

import sys as _sys
_sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from config.database import get_db
from config.security import (
    RateLimiter, generate_csrf_token, escape_html, audit_log
)
from config import storage
from admin.database import log_stock_change, deduct_order_stock, restore_order_stock, _sync_product_total_stock

ROUTER_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(ROUTER_DIR)
ADMIN_DIR = os.path.join(PROJECT_ROOT, 'admin')
SESSIONS_FILE = os.path.join(ADMIN_DIR, '.sessions.json')
CSRF_FILE = os.path.join(ADMIN_DIR, '.csrf_tokens.json')


def _get_build_version():
    try:
        r = subprocess.run(['git', 'rev-parse', '--short', 'HEAD'],
                           capture_output=True, text=True, cwd=str(PROJECT_ROOT), timeout=3)
        if r.returncode == 0 and r.stdout.strip():
            return r.stdout.strip()
    except Exception:
        pass
    return 'dev'


ADMIN_USERNAME = os.environ.get('ADMIN_USERNAME', 'admin')
ADMIN_PASSWORD_HASH = os.environ.get('ADMIN_PASSWORD_HASH', '')
ADMIN_PASSWORD_SALT = os.environ.get('ADMIN_PASSWORD_SALT', '')
CORS_ORIGIN = os.environ.get('CORS_ORIGIN', 'https://adalina-v2.onrender.com')

_login_limiter = RateLimiter()

DEFAULT_PASSWORD_HASH = '240be518fabd2724ddb6f04eeb1da5967448d7e831c08c8fa822809f74c720a9'

PBKDF2_ITERATIONS = 600000

ALLOWED_PRODUCT_COLUMNS = {'name', 'description', 'price', 'sale_price', 'category_id', 'image', 'images', 'badge', 'sizes', 'colors', 'stock', 'brand', 'rating', 'featured', 'new_arrival', 'status'}
ALLOWED_CATEGORY_COLUMNS = {'name', 'slug', 'description', 'image', 'status', 'size_system'}
ALLOWED_COLLECTION_COLUMNS = {'name', 'description', 'image', 'status'}
ALLOWED_ORDER_COLUMNS = {'status', 'total', 'items', 'customer_name', 'customer_phone', 'wilaya', 'commune', 'shipping_address', 'payment_method', 'delivery_fee', 'customer_id', 'is_read'}
ALLOWED_CUSTOMER_COLUMNS = {'name', 'email', 'phone', 'address', 'status'}


def _hash_password(password, salt='', iterations=PBKDF2_ITERATIONS):
    return hashlib.pbkdf2_hmac('sha256', password.encode(), salt.encode(), iterations).hex()


if not ADMIN_PASSWORD_HASH or ADMIN_PASSWORD_HASH == DEFAULT_PASSWORD_HASH:
    print("[WARNING] ADMIN_PASSWORD_HASH is not set or is the default.")
    print("[WARNING] Set ADMIN_PASSWORD_HASH env var to a secure hash. Generating a random one for this session.")
    ADMIN_PASSWORD_SALT = secrets.token_hex(16)
    ADMIN_PASSWORD_HASH = _hash_password('daiaaadmin02', ADMIN_PASSWORD_SALT)
    print(f"[WARNING] For this session only, login with: admin / daiaaadmin02")
    print(f"[WARNING] Set ADMIN_PASSWORD_HASH and ADMIN_PASSWORD_SALT in Render env vars.")


def _verify_password(password, stored_hash, salt=''):
    new_hash = _hash_password(password, salt, PBKDF2_ITERATIONS)
    if _hmac.compare_digest(new_hash, stored_hash):
        return True
    old_hash = _hash_password(password, salt, 100000)
    return _hmac.compare_digest(old_hash, stored_hash)


def load_sessions():
    try:
        with open(SESSIONS_FILE, 'r') as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return {}


def save_sessions(sessions):
    with open(SESSIONS_FILE, 'w') as f:
        fcntl.flock(f, fcntl.LOCK_EX)
        json.dump(sessions, f)
        fcntl.flock(f, fcntl.LOCK_UN)


def get_session(token):
    sessions = load_sessions()
    data = sessions.get(token)
    if data and data.get('admin_logged_in'):
        created = data.get('created_at', 0)
        is_remember = data.get('remember', False)
        max_age = 30 * 24 * 3600 if is_remember else 24 * 3600
        if time.time() - created > max_age:
            sessions.pop(token, None)
            save_sessions(sessions)
            return None
        return data
    return None


def create_session(remember=False):
    token = secrets.token_hex(32)
    sessions = load_sessions()
    sessions[token] = {
        'admin_logged_in': True,
        'admin_username': ADMIN_USERNAME,
        'created_at': time.time(),
        'last_active': time.time(),
        'remember': remember
    }
    save_sessions(sessions)
    return token


def delete_session(token):
    sessions = load_sessions()
    sessions.pop(token, None)
    save_sessions(sessions)


def save_csrf_token(session_token, csrf_token):
    try:
        with open(CSRF_FILE, 'r') as f:
            tokens = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        tokens = {}
    tokens[session_token] = csrf_token
    with open(CSRF_FILE, 'w') as f:
        json.dump(tokens, f)


def get_csrf_token_for_session(session_token):
    try:
        with open(CSRF_FILE, 'r') as f:
            tokens = json.load(f)
        return tokens.get(session_token)
    except (FileNotFoundError, json.JSONDecodeError):
        return None


def touch_session(token):
    sessions = load_sessions()
    if token in sessions:
        sessions[token]['last_active'] = time.time()
        save_sessions(sessions)


def get_token_from_cookies(cookie_header):
    if not cookie_header:
        return None
    cookies = http.cookies.SimpleCookie(cookie_header)
    tc = cookies.get('admin_session')
    return tc.value if tc else None


def _filter_columns(data, allowed):
    return {k: v for k, v in data.items() if k in allowed}


def _signal_cache_invalidate():
    try:
        signal_path = os.path.join(PROJECT_ROOT, '.cache_invalidate')
        with open(signal_path, 'w') as f:
            f.write(str(time.time()))
    except Exception:
        pass


def secure_path(base_dir, requested_path):
    real = os.path.realpath(os.path.join(base_dir, requested_path))
    if not real.startswith(os.path.realpath(base_dir)):
        return None
    return real


class _SafeJSONResponse(_StarletteJSONResponse):
    def render(self, content) -> bytes:
        import datetime as _dt
        def _default(o):
            if isinstance(o, (_dt.datetime, _dt.date, _dt.time)):
                return o.isoformat()
            raise TypeError(f'Object of type {type(o).__name__} is not JSON serializable')
        return json.dumps(content, default=_default, ensure_ascii=False, allow_nan=False, indent=None, separators=(',', ':')).encode('utf-8')


def _json_response(data, status_code=200):
    return _SafeJSONResponse(content=data, status_code=status_code)


def row_to_dict(row):
    return dict(row) if row else None


def rows_to_list(rows):
    return [dict(r) for r in rows]


def _parse_product_fields(d):
    if d.get('images') and isinstance(d['images'], str):
        d['images'] = json.loads(d['images'])
    if d.get('sizes') and isinstance(d['sizes'], str):
        d['sizes'] = json.loads(d['sizes'])
    if d.get('colors') and isinstance(d['colors'], str):
        d['colors'] = json.loads(d['colors'])
    if not d.get('image') and d.get('images') and isinstance(d['images'], list) and len(d['images']) > 0:
        d['image'] = d['images'][0]
    d['featured'] = bool(d.get('featured', 0))
    d['new_arrival'] = bool(d.get('new_arrival', 0))
    d['category'] = d.get('category_name') or ''


def batch_enrich_products(rows, cur):
    if not rows:
        return []
    product_ids = [dict(r)['id'] for r in rows]
    cur.execute("SELECT id, product_id, color_name, color_hex, sku, stock FROM product_variants WHERE product_id=ANY(%s) ORDER BY sort_order, id", (product_ids,))
    all_variants = cur.fetchall()
    if not all_variants:
        product_ids_str = ','.join(['%s'] * len(product_ids))
        cur.execute(f"SELECT product_id, color_name, color_hex, stock FROM product_colors WHERE product_id IN ({product_ids_str}) ORDER BY id", product_ids)
        pc_map = {}
        for pc in cur.fetchall():
            pc_map.setdefault(pc['product_id'], []).append(pc)
        result = []
        for r in rows:
            d = dict(r)
            _parse_product_fields(d)
            d['variants'] = []
            pid = d['id']
            pcs = pc_map.get(pid, [])
            if pcs:
                d['colors'] = [{'name': pc['color_name'], 'hex': pc['color_hex'], 'stock': pc['stock']} for pc in pcs]
            result.append(d)
        return result
    variant_ids = [v['id'] for v in all_variants]
    cur.execute("SELECT variant_id, image_path FROM variant_images WHERE variant_id=ANY(%s) ORDER BY sort_order", (variant_ids,))
    vi_map = {}
    for vi in cur.fetchall():
        vi_map.setdefault(vi['variant_id'], []).append(vi['image_path'])
    cur.execute("SELECT variant_id, size_name, stock, COALESCE(sku, '') AS sku FROM variant_sizes WHERE variant_id=ANY(%s) ORDER BY id", (variant_ids,))
    vs_map = {}
    for vs in cur.fetchall():
        vs_map.setdefault(vs['variant_id'], []).append(vs)
    pv_map = {}
    for v in all_variants:
        pv_map.setdefault(v['product_id'], []).append(v)
    result = []
    for r in rows:
        d = dict(r)
        _parse_product_fields(d)
        pid = d['id']
        variant_rows = pv_map.get(pid, [])
        if variant_rows:
            variants = []
            all_colors = {}
            all_sizes = set()
            images_seen = set()
            merged_sizes = []
            for v in variant_rows:
                v_images = vi_map.get(v['id'], [])
                v_sizes = [{'size': s['size_name'], 'stock': s['stock'], 'sku': s['sku']} for s in vs_map.get(v['id'], [])]
                vdict = {'id': v['id'], 'color_name': v['color_name'], 'color_hex': v['color_hex'], 'sku': v['sku'], 'stock': v['stock'], 'images': v_images, 'sizes': v_sizes}
                variants.append(vdict)
                if v['color_name'] and v['color_name'] not in all_colors:
                    all_colors[v['color_name']] = {'name': v['color_name'], 'hex': v['color_hex'], 'stock': v['stock']}
                for s in v_sizes:
                    if s['size'] not in all_sizes:
                        all_sizes.add(s['size'])
                        merged_sizes.append(s)
                for img in v_images:
                    if img not in images_seen:
                        images_seen.add(img)
            d['colors'] = list(all_colors.values()) if all_colors else []
            d['sizes'] = merged_sizes if merged_sizes else []
            d['variants'] = variants
            d['images'] = list(images_seen) if images_seen else (d.get('images') or [])
        else:
            d['variants'] = []
        result.append(d)
    return result


def enrich_product(row, cur):
    return batch_enrich_products([row], cur)[0]


def calculate_order_risk(order, cur):
    score = 0
    reasons = []
    phone = order.get('customer_phone', '')
    cur.execute("SELECT COUNT(*) AS cnt FROM orders WHERE customer_phone = %s AND id != %s", (phone, order['id']))
    prev_orders = cur.fetchone()['cnt']
    if prev_orders == 0:
        score += 30
        reasons.append('Nouveau client')
    if (order.get('total') or 0) > 20000:
        score += 25
        reasons.append('Montant élevé')
    wilaya = (order.get('wilaya') or '').lower()
    if phone and len(phone) >= 2:
        prefix = phone[:2]
        if prefix in ('05', '06', '07') and 'alger' not in wilaya:
            score += 20
            reasons.append('Préfixe téléphone ≠ wilaya')
        elif prefix in ('34', '55', '66', '77') and 'oran' not in wilaya and 'alg' not in wilaya:
            score += 20
            reasons.append('Préfixe téléphone ≠ wilaya')
    if order.get('customer_id'):
        cur.execute("SELECT created_at FROM customers WHERE id = %s", (order['customer_id'],))
        cust = cur.fetchone()
        if cust and order.get('created_at'):
            diff = (order['created_at'] - cust['created_at']).total_seconds()
            if diff < 300:
                score += 10
                reasons.append('Commande rapide après inscription')
    return min(score, 100), reasons


def _get_client_ip(request: Request) -> str:
    real_ip = request.headers.get('x-real-for', '')
    if real_ip:
        return real_ip.split(',')[0].strip()
    forwarded = request.headers.get('x-forwarded-for', '')
    if forwarded:
        return forwarded.split(',')[0].strip()
    if request.client:
        return request.client.host
    return '0.0.0.0'


def require_admin_auth(request: Request) -> str:
    token = request.cookies.get('admin_session')
    if not token or not get_session(token):
        raise HTTPException(status_code=302, headers={'Location': '/gestion/login'})
    touch_session(token)
    return token


def validate_admin_csrf(request: Request):
    token = request.cookies.get('admin_session')
    if not token or not get_session(token):
        return
    csrf_cookie = request.cookies.get('csrf_token', '')
    csrf_header = request.headers.get('X-CSRF-Token', '')
    if not csrf_cookie or not csrf_header or not secrets.compare_digest(csrf_header, csrf_cookie):
        raise HTTPException(status_code=403, detail='CSRF validation failed')


router = APIRouter(prefix='/api', tags=['admin-api'])


# ---------------------------------------------------------------------------
# GET routes
# ---------------------------------------------------------------------------




@router.get('/dashboard/stats')
def dashboard_stats(request: Request, session_token: str = Depends(require_admin_auth)):
    db = get_db()
    cur = db.cursor()
    try:
        cur.execute("""
            SELECT
                COALESCE(SUM(CASE WHEN status='delivered' THEN total ELSE 0 END), 0) AS revenue,
                COUNT(*) AS orders_count
            FROM orders
        """)
        stats_row = cur.fetchone()
        revenue = stats_row['revenue']
        orders_count = stats_row['orders_count']

        cur.execute("SELECT COUNT(*) AS cnt FROM customers")
        customers_count = cur.fetchone()['cnt']
        cur.execute("SELECT COUNT(*) AS cnt FROM products")
        products_count = cur.fetchone()['cnt']

        cur.execute("""
            SELECT o.id, o.order_number, COALESCE(c.name, o.customer_name) AS customer_name,
                   o.status, o.total, o.created_at
            FROM orders o LEFT JOIN customers c ON o.customer_id = c.id
            ORDER BY o.id DESC LIMIT 5
        """)
        recent_orders = cur.fetchall()

        cur.execute("SELECT id, name, price, image, stock FROM products ORDER BY id DESC LIMIT 5")
        recent_products = cur.fetchall()

        cur.execute("""
            WITH order_items AS (
                SELECT (item->>'product_id')::int AS product_id,
                       COALESCE((item->>'quantity')::int, 1) AS quantity
                FROM orders o,
                     jsonb_array_elements(o.items::jsonb) AS item
                WHERE o.items IS NOT NULL AND o.items != '[]'
            )
            SELECT p.id, p.name, p.price, p.image, p.stock, p.status,
                   COALESCE(SUM(oi.quantity), 0) AS sold
            FROM products p
            LEFT JOIN order_items oi ON oi.product_id = p.id
            GROUP BY p.id, p.name, p.price, p.image, p.stock, p.status
            ORDER BY sold DESC
        """)
        all_products_data = [dict(r) for r in cur.fetchall()]
        all_products_data = batch_enrich_products(all_products_data, cur)
        for p in all_products_data:
            if not p.get('image') and p.get('images') and isinstance(p['images'], list) and len(p['images']) > 0:
                p['image'] = p['images'][0]
        top_products_data = all_products_data[:5]
        unsold_products_data = [p for p in all_products_data if p['sold'] == 0][:10]

        try:
            cur.execute("""
                SELECT
                    COUNT(*) FILTER (WHERE quantity > 0 AND quantity <= low_stock_threshold) AS low_stock,
                    COUNT(*) FILTER (WHERE quantity = 0) AS out_of_stock
                FROM inventory
            """)
            stock_row = cur.fetchone()
            low_stock = stock_row['low_stock']
            out_of_stock = stock_row['out_of_stock']
        except Exception:
            low_stock = 0
            out_of_stock = 0

        cur.execute("""
            SELECT EXTRACT(MONTH FROM created_at) AS month, COUNT(*) AS cnt, COALESCE(SUM(total),0) AS rev
            FROM orders WHERE status='delivered' AND EXTRACT(YEAR FROM created_at) = EXTRACT(YEAR FROM NOW())
            GROUP BY month ORDER BY month
        """)
        monthly_data = cur.fetchall()
        monthly_orders = [0]*12
        monthly_revenue = [0]*12
        for row in monthly_data:
            m = int(row['month']) - 1
            monthly_orders[m] = row['cnt']
            monthly_revenue[m] = round(row['rev'], 2)

        most_sold_chart = []
        for p in top_products_data[:5]:
            if p['sold'] > 0:
                most_sold_chart.append({'name': p['name'], 'sold': p['sold']})

        cur.execute("SELECT COUNT(*) AS total, COUNT(*) FILTER (WHERE stock > 5) AS healthy FROM products WHERE status='active'")
        stock_health = cur.fetchone()
        stock_pct = (stock_health['healthy'] / max(stock_health['total'], 1)) * 100

        cur.execute("SELECT COUNT(*) AS cnt FROM orders WHERE status IN ('new', 'pending', 'confirmed', 'preparing') AND created_at < NOW() - INTERVAL '24 hours'")
        backlog_count = cur.fetchone()['cnt']
        backlog_pct = max(0, 100 - (backlog_count * 20))

        cur.execute("SELECT COALESCE(SUM(total), 0) AS rev FROM orders WHERE status='delivered' AND created_at >= date_trunc('month', NOW())")
        current_month_rev = cur.fetchone()['rev']
        cur.execute("SELECT COALESCE(SUM(total), 0) AS rev FROM orders WHERE status='delivered' AND created_at >= date_trunc('month', NOW() - INTERVAL '1 month') AND created_at < date_trunc('month', NOW())")
        prev_month_rev = cur.fetchone()['rev']
        target = prev_month_rev * 1.1
        rev_pct = min(100, (current_month_rev / max(target, 1)) * 100)

        health_score = int(stock_pct * 0.3 + backlog_pct * 0.25 + rev_pct * 0.25 + 80 * 0.2)
        health_score = min(100, max(0, health_score))

        cur.execute("""
            SELECT EXTRACT(DOW FROM created_at) AS dow,
                   COUNT(*) AS order_count,
                   COALESCE(SUM(total), 0) AS revenue
            FROM orders
            WHERE created_at >= NOW() - INTERVAL '7 days'
            GROUP BY EXTRACT(DOW FROM created_at)
        """)
        weekly_orders = {int(r['dow']): {'count': r['order_count'], 'revenue': float(r['revenue'])} for r in cur.fetchall()}

        try:
            cur.execute("""
                SELECT EXTRACT(DOW FROM created_at) AS dow,
                       COUNT(*) AS views
                FROM search_events
                WHERE event_type = 'page_view' AND created_at >= NOW() - INTERVAL '7 days'
                GROUP BY EXTRACT(DOW FROM created_at)
            """)
            weekly_views = {int(r['dow']): r['views'] for r in cur.fetchall()}
        except Exception:
            weekly_views = {}

        try:
            cur.execute("""
                SELECT COUNT(*) AS orders_today,
                       COALESCE(SUM(total), 0) AS revenue_today
                FROM orders
                WHERE DATE(created_at) = CURRENT_DATE
            """)
            today_stats = cur.fetchone()
        except Exception:
            today_stats = {'orders_today': 0, 'revenue_today': 0}

        return {
            'revenue': round(revenue, 2),
            'orders_count': orders_count,
            'customers_count': customers_count,
            'products_count': products_count,
            'low_stock': low_stock,
            'out_of_stock': out_of_stock,
            'recent_orders': rows_to_list(recent_orders),
            'recent_products': rows_to_list(recent_products),
            'top_products': top_products_data,
            'unsold_products': unsold_products_data,
            'monthly_orders': monthly_orders,
            'monthly_revenue': monthly_revenue,
            'most_sold_chart': most_sold_chart,
            'health_score': health_score,
            'backlog_count': backlog_count,
            'weekly_orders': weekly_orders,
            'weekly_views': weekly_views,
            'orders_today': today_stats['orders_today'],
            'revenue_today': float(today_stats['revenue_today']),
        }
    finally:
        try:
            cur.close()
        except Exception:
            pass
        try:
            db.close()
        except Exception:
            pass


@router.get('/analytics')
def analytics(request: Request, session_token: str = Depends(require_admin_auth)):
    db = get_db()
    cur = db.cursor()
    try:
        cur.execute("""
            SELECT
                COALESCE(SUM(CASE WHEN status='delivered' THEN total ELSE 0 END), 0) AS total_revenue,
                COUNT(*) AS total_orders,
                COALESCE(AVG(total), 0) AS avg_order
            FROM orders
        """)
        agg = cur.fetchone()
        total_revenue = agg['total_revenue']
        total_orders = agg['total_orders']
        avg_order = agg['avg_order']
        cur.execute("SELECT COUNT(*) AS cnt FROM customers")
        total_customers = cur.fetchone()['cnt']
        cur.execute("SELECT COUNT(*) AS cnt FROM products")
        total_products = cur.fetchone()['cnt']

        cur.execute("""
            SELECT TO_CHAR(created_at, 'YYYY-MM') AS month,
                   COUNT(*) AS orders,
                   COALESCE(SUM(total),0) AS revenue
            FROM orders
            WHERE created_at >= NOW() - INTERVAL '12 months'
            GROUP BY month ORDER BY month
        """)
        monthly_sales = cur.fetchall()

        cur.execute("""
            WITH order_items AS (
                SELECT o.id AS order_id,
                       (item->>'product_id')::int AS product_id,
                       COALESCE((item->>'quantity')::int, 1) AS quantity,
                       COALESCE((item->>'price')::numeric, 0) AS price
                FROM orders o,
                     jsonb_array_elements(o.items::jsonb) AS item
                WHERE o.items IS NOT NULL AND o.items != '[]'
            )
            SELECT c.id, c.name,
                   COUNT(DISTINCT oi.order_id) AS order_count,
                   COALESCE(SUM(oi.quantity * oi.price), 0) AS revenue
            FROM categories c
            JOIN products p ON p.category_id = c.id
            JOIN order_items oi ON oi.product_id = p.id
            GROUP BY c.id, c.name
            ORDER BY revenue DESC
        """)
        category_perf = [dict(r) for r in cur.fetchall()]

        cur.execute("""
            WITH order_items AS (
                SELECT (item->>'product_id')::int AS product_id,
                       COALESCE((item->>'quantity')::int, 1) AS quantity,
                       COALESCE((item->>'price')::numeric, 0) AS price
                FROM orders o,
                     jsonb_array_elements(o.items::jsonb) AS item
                WHERE o.items IS NOT NULL AND o.items != '[]'
            )
            SELECT p.id, p.name, p.price, p.image, p.stock,
                   COALESCE(SUM(oi.quantity), 0) AS sold,
                   COALESCE(SUM(oi.quantity * oi.price), 0) AS revenue
            FROM products p
            LEFT JOIN order_items oi ON oi.product_id = p.id
            GROUP BY p.id, p.name, p.price, p.image, p.stock
            ORDER BY sold DESC
            LIMIT 10
        """)
        best_sellers_list = [dict(r) for r in cur.fetchall()]
        for item in best_sellers_list:
            item['revenue'] = round(float(item['revenue']), 2)

        cur.execute("""
            SELECT DATE(created_at) AS day,
                   COUNT(*) AS orders,
                   COALESCE(SUM(total),0) AS revenue
            FROM orders
            WHERE created_at >= CURRENT_DATE - INTERVAL '30 days'
            GROUP BY day ORDER BY day
        """)
        daily_sales = cur.fetchall()

        conversion = (total_orders / max(total_customers, 1)) * 100

        return {
            'stats': {
                'revenue': round(total_revenue, 2),
                'orders': total_orders,
                'customers': total_customers,
                'products': total_products,
                'avg_order_value': round(avg_order, 2),
                'conversion': round(conversion, 1),
            },
            'monthly_sales': rows_to_list(monthly_sales),
            'category_performance': category_perf,
            'best_sellers': best_sellers_list,
            'daily_sales': rows_to_list(daily_sales),
        }
    finally:
        try:
            cur.close()
        except Exception:
            pass
        try:
            db.close()
        except Exception:
            pass


@router.get('/products')
def list_products(request: Request, search: str = Query(''), category: str = Query(''), session_token: str = Depends(require_admin_auth)):
    db = get_db()
    cur = db.cursor()
    try:
        params = []
        where = []
        if category:
            where.append("LOWER(c.name) = LOWER(%s)")
            params.append(category)
        if search:
            where.append("LOWER(p.name) LIKE %s")
            params.append(f'%{search.strip().lower()}%')
        where_clause = " AND ".join(where) if where else "1=1"
        cur.execute("""SELECT p.*, c.name AS category_name FROM products p
                 LEFT JOIN categories c ON p.category_id = c.id
                 WHERE """ + where_clause + " ORDER BY p.id", params)
        rows = cur.fetchall()
        result = batch_enrich_products(rows, cur)
        return result
    finally:
        try:
            cur.close()
        except Exception:
            pass
        try:
            db.close()
        except Exception:
            pass


@router.get('/products/{pid}')
def get_product(pid: str, session_token: str = Depends(require_admin_auth)):
    db = get_db()
    cur = db.cursor()
    try:
        cur.execute("""SELECT p.*, c.name AS category_name FROM products p
                             LEFT JOIN categories c ON p.category_id = c.id WHERE p.id=%s""", (pid,))
        row = cur.fetchone()
        if not row:
            return _json_response({'error': 'Not found'}, 404)
        data = dict(row)
        if data.get('images') and isinstance(data['images'], str):
            data['images'] = json.loads(data['images'])
        cur.execute("SELECT id, color_name, color_hex, sku FROM product_variants WHERE product_id=%s ORDER BY sort_order, id", (pid,))
        variant_rows = cur.fetchall()
        if variant_rows:
            variant_ids = [v['id'] for v in variant_rows]
            cur.execute("SELECT variant_id, image_path FROM variant_images WHERE variant_id=ANY(%s) ORDER BY sort_order", (variant_ids,))
            vi_map = {}
            for vi in cur.fetchall():
                vi_map.setdefault(vi['variant_id'], []).append(vi['image_path'])
            cur.execute("SELECT variant_id, size_name, stock, COALESCE(sku, '') AS sku FROM variant_sizes WHERE variant_id=ANY(%s) ORDER BY id", (variant_ids,))
            vs_map = {}
            for vs in cur.fetchall():
                vs_map.setdefault(vs['variant_id'], []).append({'size': vs['size_name'], 'stock': vs['stock'], 'sku': vs.get('sku', '')})
            variants = []
            all_colors = []
            for v in variant_rows:
                vid = v['id']
                variants.append({
                    'color_name': v['color_name'],
                    'color_hex': v['color_hex'],
                    'sku': v['sku'],
                    'images': vi_map.get(vid, []),
                    'sizes': vs_map.get(vid, []),
                })
                all_colors.append({'name': v['color_name'], 'hex': v['color_hex']})
            data['colors'] = all_colors
            data['variants'] = variants
        else:
            cur.execute("SELECT size, stock FROM product_sizes WHERE product_id=%s ORDER BY id", (pid,))
            data['sizes'] = [{'size': r['size'], 'stock': r['stock']} for r in cur.fetchall()]
            cur.execute("SELECT color_name, color_hex, stock FROM product_colors WHERE product_id=%s ORDER BY id", (pid,))
            data['colors'] = [{'name': r['color_name'], 'hex': r['color_hex'], 'stock': r['stock']} for r in cur.fetchall()]
            cur.execute("SELECT color_name, size_name, stock FROM product_variants WHERE product_id=%s ORDER BY id", (pid,))
            data['variants'] = [{'color_name': r['color_name'], 'size_name': r['size_name'], 'stock': r['stock']} for r in cur.fetchall()]
        return data
    finally:
        try:
            cur.close()
        except Exception:
            pass
        try:
            db.close()
        except Exception:
            pass


@router.get('/categories')
def list_categories(session_token: str = Depends(require_admin_auth)):
    db = get_db()
    cur = db.cursor()
    try:
        cur.execute("""
            SELECT c.*, COUNT(p.id) AS product_count
            FROM categories c
            LEFT JOIN products p ON p.category_id = c.id
            GROUP BY c.id
            ORDER BY c.id
        """)
        rows = cur.fetchall()
        return rows_to_list(rows)
    finally:
        try:
            cur.close()
        except Exception:
            pass
        try:
            db.close()
        except Exception:
            pass


@router.get('/categories/{cid}')
def get_category(cid: str, session_token: str = Depends(require_admin_auth)):
    db = get_db()
    cur = db.cursor()
    try:
        cur.execute("""
            SELECT c.*, COUNT(p.id) AS product_count
            FROM categories c
            LEFT JOIN products p ON p.category_id = c.id
            WHERE c.id=%s
            GROUP BY c.id
        """, (cid,))
        row = cur.fetchone()
        if not row:
            return _json_response({'error': 'Not found'}, 404)
        return row_to_dict(row)
    finally:
        try:
            cur.close()
        except Exception:
            pass
        try:
            db.close()
        except Exception:
            pass


@router.get('/collections/all')
def list_all_collections_and_products(session_token: str = Depends(require_admin_auth)):
    db = get_db()
    cur = db.cursor()
    try:
        cur.execute("SELECT * FROM collections ORDER BY id")
        collections = rows_to_list(cur.fetchall())
        cur.execute("SELECT id, name FROM products ORDER BY name")
        products = rows_to_list(cur.fetchall())
        return {'collections': collections, 'products': products}
    finally:
        try:
            cur.close()
        except Exception:
            pass
        try:
            db.close()
        except Exception:
            pass


@router.get('/collections')
def list_collections(session_token: str = Depends(require_admin_auth)):
    db = get_db()
    cur = db.cursor()
    try:
        cur.execute("""
            SELECT c.*, COUNT(cp.product_id) AS product_count
            FROM collections c
            LEFT JOIN collection_products cp ON cp.collection_id = c.id
            GROUP BY c.id
            ORDER BY c.id
        """)
        rows = cur.fetchall()
        return rows_to_list(rows)
    finally:
        try:
            cur.close()
        except Exception:
            pass
        try:
            db.close()
        except Exception:
            pass


@router.get('/collections/{cid}')
def get_collection(cid: str, session_token: str = Depends(require_admin_auth)):
    db = get_db()
    cur = db.cursor()
    try:
        cur.execute("SELECT * FROM collections WHERE id=%s", (cid,))
        row = cur.fetchone()
        if not row:
            return _json_response({'error': 'Not found'}, 404)
        coll = dict(row)
        cur.execute("""
            SELECT p.id, p.name, p.price, p.image FROM products p
            JOIN collection_products cp ON cp.product_id = p.id
            WHERE cp.collection_id = %s
        """, (cid,))
        prods = cur.fetchall()
        coll['product_ids'] = [p['id'] for p in prods]
        coll['products'] = rows_to_list(prods)
        return coll
    finally:
        try:
            cur.close()
        except Exception:
            pass
        try:
            db.close()
        except Exception:
            pass


@router.get('/orders')
def list_orders(session_token: str = Depends(require_admin_auth)):
    db = get_db()
    cur = db.cursor()
    try:
        cur.execute("""
            SELECT o.*, COALESCE(c.name, o.customer_name) AS customer_name, c.email AS customer_email
            FROM orders o LEFT JOIN customers c ON o.customer_id = c.id ORDER BY o.id DESC
        """)
        rows = cur.fetchall()
        orders_list = rows_to_list(rows)
        for o in orders_list:
            risk_score, risk_reasons = calculate_order_risk(o, cur)
            o['risk_score'] = risk_score
            o['risk_reasons'] = risk_reasons
        return orders_list
    finally:
        try:
            cur.close()
        except Exception:
            pass
        try:
            db.close()
        except Exception:
            pass


@router.get('/orders/{oid}/history')
def order_history(oid: str, session_token: str = Depends(require_admin_auth)):
    db = get_db()
    cur = db.cursor()
    try:
        cur.execute("SELECT status, note, created_at FROM status_history WHERE order_id=%s ORDER BY created_at ASC", (oid,))
        rows = cur.fetchall()
        result = [{'status': r['status'], 'note': r['note'], 'created_at': r['created_at'].isoformat() if r['created_at'] else None} for r in rows]
        return result
    finally:
        try:
            cur.close()
        except Exception:
            pass
        try:
            db.close()
        except Exception:
            pass


@router.get('/orders/export/csv')
def export_orders_csv(request: Request, session_token: str = Depends(require_admin_auth)):
    from starlette.responses import StreamingResponse
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    db = get_db()
    cur = db.cursor()
    try:
        cur.execute("""
            SELECT o.id, o.order_number, COALESCE(c.name, o.customer_name) AS customer_name,
                   o.customer_phone, o.wilaya, o.commune, o.total, o.status,
                   o.delivery_mode, o.delivery_fee, o.payment_method, o.items,
                   o.created_at
            FROM orders o LEFT JOIN customers c ON o.customer_id = c.id
            ORDER BY o.created_at DESC
        """)
        rows = cur.fetchall()
        wb = Workbook()
        ws = wb.active
        ws.title = 'Commandes'
        headers = ['Numéro', 'Client', 'Téléphone', 'Wilaya', 'Commune',
                   'Total (DA)', 'Statut', 'Livraison', 'Frais livraison',
                   'Paiement', 'Articles', 'Date']
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='1B4D3E', end_color='1B4D3E', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        for ri, r in enumerate(rows, 2):
            rd = dict(r)
            items_str = ''
            try:
                items = json.loads(rd['items']) if rd['items'] else []
                parts = []
                for it in items:
                    name = it.get('name') or it.get('product_name', '')
                    sz = it.get('size') or it.get('selectedSize', '')
                    qty = it.get('quantity') or it.get('qty', 1)
                    parts.append(f"{name} ({sz}) x{qty}")
                items_str = '; '.join(parts)
            except Exception:
                items_str = str(rd.get('items', ''))
            created = rd['created_at'].strftime('%Y-%m-%d %H:%M') if rd['created_at'] else ''
            vals = [rd.get('order_number', ''), rd.get('customer_name', ''),
                    rd.get('customer_phone', ''), rd.get('wilaya', ''), rd.get('commune', ''),
                    rd.get('total', 0), rd.get('status', ''), rd.get('delivery_mode', ''),
                    rd.get('delivery_fee', 0), rd.get('payment_method', ''), items_str, created]
            for col, v in enumerate(vals, 1):
                cell = ws.cell(row=ri, column=col, value=v)
                cell.border = thin_border
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 35)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            iter([buf.getvalue()]),
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': 'attachment; filename="commandes_adalina.xlsx"'}
        )
    finally:
        try:
            cur.close()
        except Exception:
            pass
        try:
            db.close()
        except Exception:
            pass


@router.get('/orders/{oid}/export/csv')
def export_order_csv(oid: str, request: Request, session_token: str = Depends(require_admin_auth)):
    from starlette.responses import StreamingResponse
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    db = get_db()
    cur = db.cursor()
    try:
        cur.execute("""
            SELECT o.id, o.order_number, COALESCE(c.name, o.customer_name) AS customer_name,
                   o.customer_phone, o.wilaya, o.commune, o.total, o.status,
                   o.delivery_mode, o.delivery_fee, o.payment_method, o.items,
                   o.created_at
            FROM orders o LEFT JOIN customers c ON o.customer_id = c.id
            WHERE o.id = %s
        """, (oid,))
        r = cur.fetchone()
        if not r:
            return _json_response({'error': 'Commande introuvable'}, 404)
        rd = dict(r)
        wb = Workbook()
        ws = wb.active
        ws.title = 'Commande'
        headers = ['Numéro', 'Client', 'Téléphone', 'Wilaya', 'Commune',
                   'Total (DA)', 'Statut', 'Livraison', 'Frais livraison',
                   'Paiement', 'Articles', 'Date']
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='1B4D3E', end_color='1B4D3E', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        items_str = ''
        try:
            items = json.loads(rd['items']) if rd['items'] else []
            parts = []
            for it in items:
                name = it.get('name') or it.get('product_name', '')
                sz = it.get('size') or it.get('selectedSize', '')
                qty = it.get('quantity') or it.get('qty', 1)
                parts.append(f"{name} ({sz}) x{qty}")
            items_str = '; '.join(parts)
        except Exception:
            items_str = str(rd.get('items', ''))
        created = rd['created_at'].strftime('%Y-%m-%d %H:%M') if rd['created_at'] else ''
        vals = [rd.get('order_number', ''), rd.get('customer_name', ''),
                rd.get('customer_phone', ''), rd.get('wilaya', ''), rd.get('commune', ''),
                rd.get('total', 0), rd.get('status', ''), rd.get('delivery_mode', ''),
                rd.get('delivery_fee', 0), rd.get('payment_method', ''), items_str, created]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=2, column=col, value=v)
            cell.border = thin_border
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 35)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        num = rd.get('order_number', oid)
        return StreamingResponse(
            iter([buf.getvalue()]),
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename="commande_{num}.xlsx"'}
        )
    finally:
        try:
            cur.close()
        except Exception:
            pass
        try:
            db.close()
        except Exception:
            pass


@router.get('/orders/{oid}')
def get_order(oid: str, session_token: str = Depends(require_admin_auth)):
    db = get_db()
    cur = db.cursor()
    try:
        cur.execute("""
            SELECT o.*, COALESCE(c.name, o.customer_name) AS customer_name, c.email AS customer_email
            FROM orders o LEFT JOIN customers c ON o.customer_id = c.id WHERE o.id=%s
        """, (oid,))
        row = cur.fetchone()
        if not row:
            return _json_response({'error': 'Not found'}, 404)
        data = dict(row)
        if data.get('items'):
            data['items'] = json.loads(data['items'])
        return data
    finally:
        try:
            cur.close()
        except Exception:
            pass
        try:
            db.close()
        except Exception:
            pass


@router.get('/settings')
def list_settings(session_token: str = Depends(require_admin_auth)):
    db = get_db()
    cur = db.cursor()
    try:
        cur.execute("SELECT * FROM settings ORDER BY category, id")
        rows = cur.fetchall()
        result = {}
        for r in rows:
            key = r['setting_key']
            val = r['setting_value']
            t = r['setting_type']
            if t == 'boolean':
                val = val == '1'
            elif t == 'number':
                try:
                    val = float(val)
                except Exception:
                    pass
            result[key] = {'value': val, 'type': t, 'category': r['category'], 'description': r.get('description', '')}
        return result
    finally:
        try:
            cur.close()
        except Exception:
            pass
        try:
            db.close()
        except Exception:
            pass


@router.get('/delivery-prices')
def list_delivery_prices(session_token: str = Depends(require_admin_auth)):
    db = get_db()
    cur = db.cursor()
    try:
        cur.execute("SELECT wilaya_id, price FROM delivery_prices ORDER BY wilaya_id")
        rows = cur.fetchall()
        result = {}
        for r in rows:
            result[str(r['wilaya_id'])] = r['price']
        return result
    finally:
        try:
            cur.close()
        except Exception:
            pass
        try:
            db.close()
        except Exception:
            pass


@router.get('/notifications')
def list_notifications(session_token: str = Depends(require_admin_auth)):
    db = get_db()
    cur = db.cursor()
    try:
        cur.execute("SELECT COUNT(*) AS cnt FROM orders WHERE is_read IS NULL OR is_read=0")
        unread = cur.fetchone()['cnt']
        cur.execute("""
            SELECT id AS order_id, order_number, customer_name, total, created_at
            FROM orders
            WHERE is_read IS NULL OR is_read=0
            ORDER BY created_at DESC LIMIT 20
        """)
        rows = cur.fetchall()
        notifs = []
        for r in rows:
            n = dict(r)
            n['id'] = n['order_id']
            n['type'] = 'order'
            n['customer_name'] = n.get('customer_name', '') or ''
            n['total'] = float(n.get('total', 0))
            notifs.append(n)
        try:
            cur.execute("""
                SELECT i.product_id, i.quantity, p.name AS product_name
                FROM inventory i
                JOIN products p ON i.product_id = p.id
                WHERE p.status = 'active' AND i.quantity <= 5 AND i.quantity >= 0
                ORDER BY i.quantity ASC LIMIT 5
            """)
            low_stock = cur.fetchall()
            for ls in low_stock:
                notifs.append({
                    'id': 'low_stock_' + str(ls['product_id']),
                    'type': 'low_stock',
                    'product_id': ls['product_id'],
                    'product_name': ls['product_name'],
                    'quantity': ls['quantity'],
                    'message': f"Stock faible ({ls['quantity']}) pour {ls['product_name']}"
                })
        except Exception:
            pass
        return {'notifications': notifs, 'unread_count': unread}
    finally:
        try:
            cur.close()
        except Exception:
            pass
        try:
            db.close()
        except Exception:
            pass


@router.get('/customers')
def list_customers(request: Request, page: int = Query(1), per_page: int = Query(20), search: str = Query(''), session_token: str = Depends(require_admin_auth)):
    db = get_db()
    cur = db.cursor()
    try:
        offset = (page - 1) * per_page
        where = []
        params = []
        if search:
            where.append("(LOWER(name) LIKE %s OR LOWER(email) LIKE %s)")
            params.extend(['%' + search.strip().lower() + '%', '%' + search.strip().lower() + '%'])
        where_clause = " WHERE " + " AND ".join(where) if where else ""
        cur.execute("SELECT COUNT(*) AS cnt FROM customers" + where_clause, params)
        total = cur.fetchone()['cnt']
        cur.execute("""
            SELECT c.*, COALESCE(oa.orders_count, 0) AS orders_count,
                   COALESCE(oa.total_spent, 0) AS total_spent
            FROM customers c
            LEFT JOIN (
                SELECT customer_id, COUNT(*) AS orders_count, COALESCE(SUM(total), 0) AS total_spent
                FROM orders WHERE customer_id IS NOT NULL GROUP BY customer_id
            ) oa ON oa.customer_id = c.id
        """ + where_clause + " ORDER BY c.id DESC LIMIT %s OFFSET %s", params + [per_page, offset])
        rows = cur.fetchall()
        total_pages = max(1, (total + per_page - 1) // per_page) if total > 0 else 1
        return {'customers': rows_to_list(rows), 'total': total, 'page': page, 'per_page': per_page, 'total_pages': total_pages}
    finally:
        try:
            cur.close()
        except Exception:
            pass
        try:
            db.close()
        except Exception:
            pass


@router.get('/customers/{cid}')
def get_customer(cid: str, session_token: str = Depends(require_admin_auth)):
    db = get_db()
    cur = db.cursor()
    try:
        cur.execute("SELECT * FROM customers WHERE id=%s", (cid,))
        row = cur.fetchone()
        if not row:
            return _json_response({'error': 'Not found'}, 404)
        cust = dict(row)
        cur.execute("SELECT * FROM orders WHERE customer_id=%s ORDER BY id DESC", (cid,))
        orders = cur.fetchall()
        cust['orders'] = rows_to_list(orders)
        return cust
    finally:
        try:
            cur.close()
        except Exception:
            pass
        try:
            db.close()
        except Exception:
            pass


@router.get('/inventory')
def list_inventory(request: Request, status: str = Query('all'), session_token: str = Depends(require_admin_auth)):
    db = get_db()
    cur = db.cursor()
    try:
        status_filter = status.strip().lower()
        base_sql = """SELECT i.*, p.name AS product_name, p.image AS product_image,
                            c.name AS category_name
                     FROM inventory i
                     JOIN products p ON i.product_id = p.id
                     LEFT JOIN categories c ON p.category_id = c.id"""
        where = []
        if status_filter == 'low':
            where.append("i.quantity > 0 AND i.quantity <= i.low_stock_threshold")
        elif status_filter == 'out':
            where.append("i.quantity = 0")
        elif status_filter == 'in':
            where.append("i.quantity > i.low_stock_threshold")
        sql = base_sql
        if where:
            sql += " WHERE " + " AND ".join(where)
        sql += " ORDER BY i.quantity ASC"
        cur.execute(sql)
        rows = cur.fetchall()
        result = rows_to_list(rows)
        for r in result:
            q = r['quantity']
            threshold = r.get('low_stock_threshold', 5)
            if q == 0:
                r['stock_status'] = 'out_of_stock'
            elif q <= threshold:
                r['stock_status'] = 'low_stock'
            else:
                r['stock_status'] = 'in_stock'
        cur.execute("""
            SELECT
                COUNT(*) AS total_count,
                COUNT(*) FILTER (WHERE quantity > low_stock_threshold) AS in_count,
                COUNT(*) FILTER (WHERE quantity > 0 AND quantity <= low_stock_threshold) AS low_count,
                COUNT(*) FILTER (WHERE quantity = 0) AS out_count
            FROM inventory
        """)
        counts_row = cur.fetchone()
        return {
            'items': result,
            'counts': {
                'total': counts_row['total_count'],
                'in_stock': counts_row['in_count'],
                'low_stock': counts_row['low_count'],
                'out_of_stock': counts_row['out_count']
            }
        }
    finally:
        try:
            cur.close()
        except Exception:
            pass
        try:
            db.close()
        except Exception:
            pass


@router.get('/inventory/{pid}/history')
def inventory_history(pid: str, session_token: str = Depends(require_admin_auth)):
    db = get_db()
    cur = db.cursor()
    try:
        cur.execute("""SELECT sh.*, p.name AS product_name
                              FROM stock_history sh
                              JOIN products p ON sh.product_id = p.id
                              WHERE sh.product_id=%s
                              ORDER BY sh.id DESC""", (pid,))
        rows = cur.fetchall()
        return rows_to_list(rows)
    finally:
        try:
            cur.close()
        except Exception:
            pass
        try:
            db.close()
        except Exception:
            pass


@router.get('/inventory/{pid}')
def get_inventory_item(pid: str, session_token: str = Depends(require_admin_auth)):
    db = get_db()
    cur = db.cursor()
    try:
        cur.execute("""SELECT i.*, p.name AS product_name, p.image AS product_image,
                                    c.name AS category_name
                             FROM inventory i
                             JOIN products p ON i.product_id = p.id
                             LEFT JOIN categories c ON p.category_id = c.id
                             WHERE i.product_id=%s""", (pid,))
        row = cur.fetchone()
        if not row:
            return _json_response({'error': 'Not found'}, 404)
        item = dict(row)
        q = item['quantity']
        threshold = item.get('low_stock_threshold', 5)
        if q == 0:
            item['stock_status'] = 'out_of_stock'
        elif q <= threshold:
            item['stock_status'] = 'low_stock'
        else:
            item['stock_status'] = 'in_stock'
        return item
    finally:
        try:
            cur.close()
        except Exception:
            pass
        try:
            db.close()
        except Exception:
            pass


@router.get('/inventory/variants/{pid}')
def get_product_variants_inventory(pid: str, session_token: str = Depends(require_admin_auth)):
    db = get_db()
    cur = db.cursor()
    try:
        cur.execute("SELECT id, color_name, color_hex, sku, stock FROM product_variants WHERE product_id=%s ORDER BY id", (pid,))
        variants = rows_to_list(cur.fetchall())
        if not variants:
            return []
        variant_ids = [v['id'] for v in variants]
        placeholders = ','.join(['%s'] * len(variant_ids))
        cur.execute(f"SELECT variant_id, size_name, stock, COALESCE(sku, '') AS sku FROM variant_sizes WHERE variant_id=ANY(%s) ORDER BY id", (variant_ids,))
        vs_rows = cur.fetchall()
        vs_map = {}
        for vs in vs_rows:
            vid = vs['variant_id']
            if vid not in vs_map:
                vs_map[vid] = []
            vs_map[vid].append({'size_name': vs['size_name'], 'stock': vs['stock'] or 0, 'sku': vs['sku']})
        for v in variants:
            v['sizes'] = vs_map.get(v['id'], [])
        return variants
    finally:
        try:
            cur.close()
        except Exception:
            pass
        try:
            db.close()
        except Exception:
            pass


@router.put('/inventory/variants/{vid}/sizes/{size}')
def update_variant_size_stock(vid: str, size: str, request: Request, data: dict = Body(...), session_token: str = Depends(require_admin_auth)):
    validate_admin_csrf(request)
    db = get_db()
    cur = db.cursor()
    try:
        stock = data.get('stock', 0)
        reason = data.get('reason', 'Manual adjustment')
        cur.execute("SELECT stock FROM variant_sizes WHERE variant_id=%s AND size_name=%s", (vid, size))
        row = cur.fetchone()
        if not row:
            return _json_response({'error': 'Variant size not found'}, 404)
        before = row['stock']
        cur.execute("UPDATE variant_sizes SET stock=%s WHERE variant_id=%s AND size_name=%s", (stock, vid, size))
        cur.execute("UPDATE product_variants SET stock = (SELECT COALESCE(SUM(stock),0) FROM variant_sizes WHERE variant_id=%s) WHERE id=%s", (vid, vid))
        cur.execute("SELECT product_id FROM product_variants WHERE id=%s", (vid,))
        pv_row = cur.fetchone()
        if pv_row:
            pid = pv_row['product_id']
            log_stock_change(cur, pid, stock - before, before, reason, int(vid), None, size)
            _sync_product_total_stock(cur, pid)
        db.commit()
        _signal_cache_invalidate()
        return {'message': 'Variant size stock updated', 'before': before, 'after': stock}
    finally:
        try:
            cur.close()
        except Exception:
            pass
        try:
            db.close()
        except Exception:
            pass


@router.get('/inventory/alerts')
def get_inventory_alerts(session_token: str = Depends(require_admin_auth)):
    db = get_db()
    cur = db.cursor()
    try:
        cur.execute("""
            SELECT vs.id AS vs_id, vs.variant_id, vs.size_name, vs.stock, vs.sku,
                   pv.color_name, pv.color_hex,
                   p.id AS product_id, p.name AS product_name, p.image AS product_image
            FROM variant_sizes vs
            JOIN product_variants pv ON vs.variant_id = pv.id
            JOIN products p ON pv.product_id = p.id
            WHERE p.status = 'active' AND vs.stock <= 5
            ORDER BY vs.stock ASC
            LIMIT 20
        """)
        rows = rows_to_list(cur.fetchall())
        for r in rows:
            if r['stock'] == 0:
                r['status'] = 'out_of_stock'
            elif r['stock'] <= 5:
                r['status'] = 'low_stock'
        return {'alerts': rows}
    finally:
        try:
            cur.close()
        except Exception:
            pass
        try:
            db.close()
        except Exception:
            pass


@router.get('/backup/download')
def backup_download(session_token: str = Depends(require_admin_auth)):
    return _json_response({'error': 'Backup via store.db is no longer available. Use pg_dump for PostgreSQL backups.'}, 400)


# ---------------------------------------------------------------------------
# POST routes
# ---------------------------------------------------------------------------

@router.post('/products')
def create_product(request: Request, data: dict = Body(...), session_token: str = Depends(require_admin_auth)):
    validate_admin_csrf(request)
    db = get_db()
    cur = db.cursor()
    try:
        cat = data.get('category_name', '')
        cat_id = None
        if cat:
            cur.execute("SELECT id FROM categories WHERE name=%s", (cat,))
            row = cur.fetchone()
            if row:
                cat_id = row['id']
        status = data.get('status', 'active')
        variants = data.get('variants', [])
        colors = data.get('colors', [])
        sizes = data.get('sizes', [])
        total_stock = data.get('stock', 0)
        cur.execute("""INSERT INTO products (name, description, price, sale_price, category_id, image, images, badge, sizes, colors, stock, brand, rating, featured, new_arrival, status, created_at)
                       VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,NOW()) RETURNING id""",
                    (data.get('name', ''), data.get('description', ''), data.get('price', 0),
                     data.get('sale_price'), cat_id, data.get('image', ''),
                     json.dumps(data.get('images', [])), data.get('badge'),
                     json.dumps(sizes), json.dumps(colors),
                     total_stock, data.get('brand', ''), data.get('rating', 0),
                     data.get('featured', 0), data.get('new_arrival', 0), status))
        pid = cur.fetchone()['id']
        color_names = [c.get('name', c) if isinstance(c, dict) else c for c in colors] if colors else []
        size_names = [s.get('size', s) if isinstance(s, dict) else s for s in sizes] if sizes else []
        for s_name in size_names:
            cur.execute("INSERT INTO product_sizes (product_id, size, stock) VALUES (%s, %s, 0)", (pid, s_name))
        for c_name in color_names:
            _hex = ''
            for c in (colors or []):
                if (c.get('name', c) if isinstance(c, dict) else c) == c_name:
                    _hex = c.get('hex', '') if isinstance(c, dict) else ''
                    break
            cur.execute("INSERT INTO product_colors (product_id, color_name, color_hex, stock) VALUES (%s, %s, %s, 0)", (pid, c_name, _hex))
        if variants and isinstance(variants, list):
            for idx, v in enumerate(variants):
                cur.execute("""INSERT INTO product_variants (product_id, color_name, color_hex, sku, sort_order, stock)
                               VALUES (%s,%s,%s,%s,%s,0) RETURNING id""",
                            (pid, v.get('color_name', ''), v.get('color_hex', ''),
                             v.get('sku', ''), idx))
                vid = cur.fetchone()['id']
                for img_idx, img_path in enumerate(v.get('images', [])):
                    if img_path:
                        cur.execute("INSERT INTO variant_images (variant_id, image_path, sort_order) VALUES (%s, %s, %s)",
                                    (vid, img_path, img_idx))
                for s in v.get('sizes', []):
                    cur.execute("INSERT INTO variant_sizes (variant_id, size_name, stock, sku) VALUES (%s, %s, %s, %s)",
                                (vid, s.get('size', s) if isinstance(s, dict) else s,
                                 s.get('stock', 0) if isinstance(s, dict) else 0,
                                 s.get('sku', '') if isinstance(s, dict) else ''))
            cur.execute("""
                SELECT COALESCE(SUM(vs.stock), 0) AS total
                FROM product_variants pv
                JOIN variant_sizes vs ON vs.variant_id = pv.id
                WHERE pv.product_id = %s
            """, (pid,))
            total_stock = cur.fetchone()['total']
            cur.execute("UPDATE products SET stock=%s WHERE id=%s", (total_stock, pid))
        cur.execute("INSERT INTO inventory (product_id, quantity) VALUES (%s, %s) ON CONFLICT (product_id) DO NOTHING", (pid, total_stock))
        db.commit()
        _signal_cache_invalidate()
        return _json_response({'id': pid, 'message': 'Product created'}, 201)
    except Exception as e:
        db.rollback()
        raise
    finally:
        try:
            cur.close()
        except Exception:
            pass
        try:
            db.close()
        except Exception:
            pass


@router.post('/categories')
def create_category(request: Request, data: dict = Body(...), session_token: str = Depends(require_admin_auth)):
    validate_admin_csrf(request)
    db = get_db()
    cur = db.cursor()
    try:
        name = data.get('name', '').strip()
        if not name:
            return _json_response({'error': 'Name required'}, 400)
        slug = data.get('slug', '') or name.lower().replace(' ', '-')
        size_system = data.get('size_system', 'standard')
        try:
            cur.execute("INSERT INTO categories (name, slug, description, image, status, size_system) VALUES (%s,%s,%s,%s,%s,%s) RETURNING id",
                        (name, slug, data.get('description', ''), data.get('image', ''), data.get('status', 'active'), size_system))
            row_id = cur.fetchone()
            db.commit()
            _signal_cache_invalidate()
            return _json_response({'id': row_id['id'] if row_id else None, 'message': 'Category created'}, 201)
        except Exception as e:
            db.rollback()
            msg = str(e)
            if 'duplicate key' in msg and 'name' in msg:
                return _json_response({'error': f'A category named "{name}" already exists'}, 409)
            elif 'duplicate key' in msg and 'slug' in msg:
                return _json_response({'error': f'A category with slug "{slug}" already exists'}, 409)
            else:
                return _json_response({'error': msg}, 500)
    finally:
        try:
            cur.close()
        except Exception:
            pass
        try:
            db.close()
        except Exception:
            pass


@router.post('/collections')
def create_collection(request: Request, data: dict = Body(...), session_token: str = Depends(require_admin_auth)):
    validate_admin_csrf(request)
    db = get_db()
    cur = db.cursor()
    try:
        cur.execute("INSERT INTO collections (name, description, image, status) VALUES (%s,%s,%s,%s) RETURNING id",
                    (data.get('name', ''), data.get('description', ''), data.get('image', ''), data.get('status', 'active')))
        cid = cur.fetchone()['id']
        for pid in data.get('product_ids', []):
            cur.execute("INSERT INTO collection_products (collection_id, product_id) VALUES (%s,%s) ON CONFLICT DO NOTHING", (cid, pid))
        db.commit()
        _signal_cache_invalidate()
        return _json_response({'id': cid, 'message': 'Collection created'}, 201)
    finally:
        try:
            cur.close()
        except Exception:
            pass
        try:
            db.close()
        except Exception:
            pass




@router.post('/customers')
def create_customer(request: Request, data: dict = Body(...), session_token: str = Depends(require_admin_auth)):
    validate_admin_csrf(request)
    db = get_db()
    cur = db.cursor()
    try:
        cur.execute("INSERT INTO customers (name, email, status) VALUES (%s,%s,%s) RETURNING id",
                    (data.get('name', ''), data.get('email', ''), data.get('status', 'active')))
        row = cur.fetchone()
        db.commit()
        _signal_cache_invalidate()
        return _json_response({'id': row['id'] if row else None, 'message': 'Customer created'}, 201)
    finally:
        try:
            cur.close()
        except Exception:
            pass
        try:
            db.close()
        except Exception:
            pass


@router.post('/inventory/{pid}/adjust')
def adjust_inventory(pid: str, request: Request, data: dict = Body(...), session_token: str = Depends(require_admin_auth)):
    validate_admin_csrf(request)
    db = get_db()
    cur = db.cursor()
    try:
        change = data.get('change', 0)
        reason = data.get('reason', '')
        variant_id = data.get('variant_id')
        size_name = data.get('size_name')
        if variant_id and size_name:
            cur.execute("SELECT stock FROM variant_sizes WHERE variant_id=%s AND size_name=%s", (variant_id, size_name))
            row = cur.fetchone()
            if not row:
                return _json_response({'error': 'Variant size not found'}, 404)
            before_qty = row['stock']
            new_qty = max(0, before_qty + change)
            cur.execute("UPDATE variant_sizes SET stock=%s WHERE variant_id=%s AND size_name=%s", (new_qty, variant_id, size_name))
            cur.execute("UPDATE product_variants SET stock = (SELECT COALESCE(SUM(stock),0) FROM variant_sizes WHERE variant_id=%s) WHERE id=%s", (variant_id, variant_id))
            log_stock_change(cur, pid, change, before_qty, reason, variant_id, None, size_name)
            _sync_product_total_stock(cur, pid)
        else:
            cur.execute("SELECT quantity FROM inventory WHERE product_id=%s", (pid,))
            before = cur.fetchone()
            before_qty = before['quantity'] if before else 0
            cur.execute("UPDATE inventory SET quantity = quantity + %s, updated_at = CURRENT_TIMESTAMP WHERE product_id=%s", (change, pid))
            cur.execute("UPDATE products SET stock = (SELECT quantity FROM inventory WHERE product_id=%s) WHERE id=%s", (pid, pid))
            log_stock_change(cur, pid, change, before_qty, reason)
        db.commit()
        _signal_cache_invalidate()
        return {'message': 'Inventory adjusted'}
    finally:
        try:
            cur.close()
        except Exception:
            pass
        try:
            db.close()
        except Exception:
            pass


@router.post('/upload')
async def upload_files(request: Request, files: List[UploadFile] = File(...), type: str = Form('products'), session_token: str = Depends(require_admin_auth)):
    MAGIC_BYTES = {
        b'\xff\xd8\xff': 'image/jpeg',
        b'\x89PNG': 'image/png',
        b'RIFF': 'image/webp',
    }
    EXT_TO_MIME = {'.jpg': 'image/jpeg', '.jpeg': 'image/jpeg', '.png': 'image/png', '.webp': 'image/webp'}
    MAX_SIZE = 10 * 1024 * 1024
    ALLOWED_EXT = set(EXT_TO_MIME.keys())
    category = type
    saved_paths = []
    errors = []
    for f in files:
        ext = os.path.splitext(f.filename or '')[1].lower()
        if ext == '.svg':
            errors.append(f'{f.filename}: SVG non autorisé')
            continue
        if ext not in ALLOWED_EXT:
            errors.append(f'{f.filename}: format non supporté')
            continue
        content = await f.read()
        if len(content) > MAX_SIZE:
            errors.append(f'{f.filename}: trop volumineux (max 10MB)')
            continue
        detected_mime = None
        for magic, mime in MAGIC_BYTES.items():
            if content[:len(magic)] == magic:
                detected_mime = mime
                break
        if detected_mime and detected_mime != EXT_TO_MIME.get(ext):
            errors.append(f'{f.filename}: extension ne correspond pas au contenu')
            continue
        if not detected_mime:
            detected_mime = EXT_TO_MIME.get(ext, 'image/jpeg')
        subdir = 'settings' if category == 'settings' else 'products'
        ts = int(time.time() * 1000)
        safe_base = re.sub(r'[^a-zA-Z0-9._-]', '_', os.path.basename(f.filename or 'upload'))
        safe_name = f"{ts}_{safe_base}"
        storage_path = f'{subdir}/{safe_name}'
        if not storage.is_enabled():
            errors.append(f'{f.filename}: storage not configured (set CLOUDINARY env vars)')
            continue
        try:
            url = storage.upload_file(content, storage_path, detected_mime)
        except Exception as ue:
            print(f"[Admin] Storage upload exception for {f.filename}: {ue}")
            url = None
        if url:
            saved_paths.append(url)
        else:
            errors.append(f'{f.filename}: upload failed')
    result = {'paths': saved_paths, 'count': len(saved_paths)}
    if errors:
        result['errors'] = errors
    return result


# ---------------------------------------------------------------------------
# PUT routes
# ---------------------------------------------------------------------------

@router.put('/products/{pid}/images/main')
def set_main_image(pid: str, request: Request, data: dict = Body(...), session_token: str = Depends(require_admin_auth)):
    validate_admin_csrf(request)
    db = get_db()
    cur = db.cursor()
    try:
        img_path = data.get('path', '')
        if img_path:
            cur.execute("SELECT images FROM products WHERE id=%s", (pid,))
            row = cur.fetchone()
            images = json.loads(row['images']) if row and row['images'] else []
            if img_path in images:
                images.remove(img_path)
                images.insert(0, img_path)
                cur.execute("UPDATE products SET images=%s, image=%s WHERE id=%s", (json.dumps(images), img_path, pid))
                db.commit()
                _signal_cache_invalidate()
                return {'message': 'Main image updated', 'images': images}
        return _json_response({'error': 'Image not found'}, 404)
    finally:
        try:
            cur.close()
        except Exception:
            pass
        try:
            db.close()
        except Exception:
            pass


@router.put('/products/{pid}')
def update_product(pid: str, request: Request, data: dict = Body(...), session_token: str = Depends(require_admin_auth)):
    validate_admin_csrf(request)
    db = get_db()
    cur = db.cursor()
    try:
        product_id = pid

        variants_data = data.pop('variants', None)

        sizes = None
        if 'sizes' in data:
            sizes = data['sizes']
            if isinstance(sizes, str):
                sizes = json.loads(sizes)
            del data['sizes']
            sizeNames = [s.get('size', s) if isinstance(s, dict) else s for s in sizes]
            data['sizes'] = json.dumps(sizeNames)

        colors = None
        if 'colors' in data:
            colors = data['colors']
            if isinstance(colors, str):
                colors = json.loads(colors)
            del data['colors']
            colorNames = [c.get('name', c) if isinstance(c, dict) else c for c in colors]
            data['colors'] = json.dumps(colorNames)

        if 'category_name' in data:
            cat = data['category_name']
            cur.execute("SELECT id FROM categories WHERE name=%s", (cat,))
            row = cur.fetchone()
            data['category_id'] = row['id'] if row else None
            del data['category_name']
        for key in ('images',):
            if key in data and isinstance(data[key], (list, dict)):
                data[key] = json.dumps(data[key])
        if 'images' in data:
            parsed = json.loads(data['images']) if isinstance(data['images'], str) else data['images']
            if isinstance(parsed, list) and len(parsed) > 0:
                data['image'] = parsed[0]

        data = _filter_columns(data, ALLOWED_PRODUCT_COLUMNS)
        if data:
            sets = ', '.join(f"{k}=%s" for k in data)
            vals = list(data.values()) + [pid]
            cur.execute(f"UPDATE products SET {sets} WHERE id=%s", vals)

        if sizes is not None:
            cur.execute("DELETE FROM product_sizes WHERE product_id=%s", (pid,))
            for s in sizes:
                sname = s.get('size', s) if isinstance(s, dict) else s
                sstock = s.get('stock', 0) if isinstance(s, dict) else 0
                cur.execute("INSERT INTO product_sizes (product_id, size, stock) VALUES (%s, %s, %s)", (pid, sname, sstock))
        if colors is not None:
            cur.execute("DELETE FROM product_colors WHERE product_id=%s", (pid,))
            for c in colors:
                cname = c.get('name', c) if isinstance(c, dict) else c
                chex = c.get('hex', '') if isinstance(c, dict) else ''
                cstock = c.get('stock', 0) if isinstance(c, dict) else 0
                cur.execute("INSERT INTO product_colors (product_id, color_name, color_hex, stock) VALUES (%s, %s, %s, %s)", (pid, cname, chex, cstock))

        computed_stock = None
        if variants_data is not None and isinstance(variants_data, list):
            is_advanced = len(variants_data) > 0 and ('images' in variants_data[0] or 'sku' in variants_data[0] or 'sizes' in variants_data[0])
            cur.execute("SELECT id FROM product_variants WHERE product_id=%s", (pid,))
            old_vids = [r['id'] for r in cur.fetchall()]
            for vid in old_vids:
                cur.execute("DELETE FROM variant_images WHERE variant_id=%s", (vid,))
                cur.execute("DELETE FROM variant_sizes WHERE variant_id=%s", (vid,))
            cur.execute("DELETE FROM product_variants WHERE product_id=%s", (pid,))
            if is_advanced:
                for idx, v in enumerate(variants_data):
                    cur.execute("""INSERT INTO product_variants (product_id, color_name, color_hex, sku, sort_order, stock)
                                   VALUES (%s,%s,%s,%s,%s,0) RETURNING id""",
                                (pid, v.get('color_name', ''), v.get('color_hex', ''),
                                 v.get('sku', ''), idx))
                    vid = cur.fetchone()['id']
                    for img_idx, img_path in enumerate(v.get('images', [])):
                        if img_path:
                            cur.execute("INSERT INTO variant_images (variant_id, image_path, sort_order) VALUES (%s, %s, %s)",
                                        (vid, img_path, img_idx))
                    for s in v.get('sizes', []):
                        cur.execute("INSERT INTO variant_sizes (variant_id, size_name, stock, sku) VALUES (%s, %s, %s, %s)",
                                    (vid, s.get('size', s) if isinstance(s, dict) else s,
                                     s.get('stock', 0) if isinstance(s, dict) else 0,
                                     s.get('sku', '') if isinstance(s, dict) else ''))
                cur.execute("""
                    SELECT COALESCE(SUM(vs.stock), 0) AS total
                    FROM product_variants pv
                    JOIN variant_sizes vs ON vs.variant_id = pv.id
                    WHERE pv.product_id = %s
                """, (pid,))
                computed_stock = cur.fetchone()['total']
                cur.execute("UPDATE products SET stock=%s WHERE id=%s", (computed_stock, pid))
            else:
                for v in variants_data:
                    cur.execute("INSERT INTO product_variants (product_id, color_name, size_name, stock) VALUES (%s, %s, %s, %s)",
                                (pid, v.get('color_name', ''), v.get('size_name', ''), v.get('stock', 0)))

        if computed_stock is not None:
            cur.execute("UPDATE inventory SET quantity=%s, updated_at=CURRENT_TIMESTAMP WHERE product_id=%s", (computed_stock, pid))
        elif 'stock' in data:
            cur.execute("UPDATE inventory SET quantity=%s, updated_at=CURRENT_TIMESTAMP WHERE product_id=%s", (data['stock'], pid))
        db.commit()
        _signal_cache_invalidate()
        return {'message': 'Product updated'}
    finally:
        try:
            cur.close()
        except Exception:
            pass
        try:
            db.close()
        except Exception:
            pass


@router.put('/categories/{cid}')
def update_category(cid: str, request: Request, data: dict = Body(...), session_token: str = Depends(require_admin_auth)):
    validate_admin_csrf(request)
    db = get_db()
    cur = db.cursor()
    try:
        data = _filter_columns(data, ALLOWED_CATEGORY_COLUMNS)
        if not data:
            return {'message': 'Category updated'}
        sets = ', '.join(f"{k}=%s" for k in data)
        vals = list(data.values()) + [cid]
        cur.execute(f"UPDATE categories SET {sets} WHERE id=%s", vals)
        db.commit()
        _signal_cache_invalidate()
        return {'message': 'Category updated'}
    finally:
        try:
            cur.close()
        except Exception:
            pass
        try:
            db.close()
        except Exception:
            pass


@router.put('/collections/{cid}/products')
def update_collection_products(cid: str, request: Request, data: dict = Body(...), session_token: str = Depends(require_admin_auth)):
    validate_admin_csrf(request)
    db = get_db()
    cur = db.cursor()
    try:
        cur.execute("DELETE FROM collection_products WHERE collection_id=%s", (cid,))
        for pid in data.get('product_ids', []):
            cur.execute("INSERT INTO collection_products (collection_id, product_id) VALUES (%s,%s) ON CONFLICT DO NOTHING", (cid, pid))
        db.commit()
        _signal_cache_invalidate()
        return {'message': 'Products updated'}
    finally:
        try:
            cur.close()
        except Exception:
            pass
        try:
            db.close()
        except Exception:
            pass


@router.put('/collections/{cid}')
def update_collection(cid: str, request: Request, data: dict = Body(...), session_token: str = Depends(require_admin_auth)):
    validate_admin_csrf(request)
    db = get_db()
    cur = db.cursor()
    try:
        data = _filter_columns(data, ALLOWED_COLLECTION_COLUMNS)
        if not data:
            return {'message': 'Collection updated'}
        sets = ', '.join(f"{k}=%s" for k in data)
        vals = list(data.values()) + [cid]
        cur.execute(f"UPDATE collections SET {sets} WHERE id=%s", vals)
        db.commit()
        _signal_cache_invalidate()
        return {'message': 'Collection updated'}
    finally:
        try:
            cur.close()
        except Exception:
            pass
        try:
            db.close()
        except Exception:
            pass


@router.put('/delivery-prices')
def update_delivery_prices(request: Request, data: dict = Body(...), session_token: str = Depends(require_admin_auth)):
    validate_admin_csrf(request)
    db = get_db()
    cur = db.cursor()
    try:
        for wilaya_id_str, price in data.items():
            wid = int(wilaya_id_str)
            p = float(price) if price else 0
            cur.execute("SELECT 1 FROM delivery_prices WHERE wilaya_id=%s", (wid,))
            if cur.fetchone():
                cur.execute("UPDATE delivery_prices SET price=%s WHERE wilaya_id=%s", (p, wid))
            else:
                cur.execute("INSERT INTO delivery_prices (wilaya_id, price) VALUES (%s, %s)", (wid, p))
        db.commit()
        _signal_cache_invalidate()
        return {'message': 'Delivery prices saved'}
    finally:
        try:
            cur.close()
        except Exception:
            pass
        try:
            db.close()
        except Exception:
            pass


@router.put('/settings')
def update_settings(request: Request, data: dict = Body(...), session_token: str = Depends(require_admin_auth)):
    validate_admin_csrf(request)
    db = get_db()
    cur = db.cursor()
    try:
        for key, val in data.items():
            if val is None:
                val = ''
            if isinstance(val, bool):
                val = '1' if val else '0'
            elif not isinstance(val, str):
                val = str(val)
            cur.execute("SELECT id FROM settings WHERE setting_key=%s", (key,))
            if cur.fetchone():
                cur.execute("UPDATE settings SET setting_value=%s, updated_at=CURRENT_TIMESTAMP WHERE setting_key=%s", (val, key))
            else:
                cur.execute("INSERT INTO settings (setting_key, setting_value, setting_type, category) VALUES (%s, %s, 'text', 'custom')", (key, val))
        db.commit()
        _signal_cache_invalidate()
        return {'message': 'Settings saved'}
    finally:
        try:
            cur.close()
        except Exception:
            pass
        try:
            db.close()
        except Exception:
            pass


@router.put('/notifications/read-all')
def mark_all_notifications_read(request: Request, session_token: str = Depends(require_admin_auth)):
    validate_admin_csrf(request)
    db = get_db()
    cur = db.cursor()
    try:
        cur.execute("UPDATE orders SET is_read=1")
        db.commit()
        _signal_cache_invalidate()
        return {'message': 'All notifications marked as read'}
    finally:
        try:
            cur.close()
        except Exception:
            pass
        try:
            db.close()
        except Exception:
            pass


@router.put('/notifications/read/{oid}')
def mark_notification_read(oid: str, request: Request, session_token: str = Depends(require_admin_auth)):
    validate_admin_csrf(request)
    db = get_db()
    cur = db.cursor()
    try:
        cur.execute("UPDATE orders SET is_read=1 WHERE id=%s", (oid,))
        db.commit()
        _signal_cache_invalidate()
        return {'message': 'Notification marked as read'}
    finally:
        try:
            cur.close()
        except Exception:
            pass
        try:
            db.close()
        except Exception:
            pass


@router.put('/orders/{oid}')
def update_order(oid: str, request: Request, data: dict = Body(...), session_token: str = Depends(require_admin_auth)):
    validate_admin_csrf(request)
    db = get_db()
    cur = db.cursor()
    try:
        old_status = None
        old_items = '[]'
        if 'status' in data:
            cur.execute("SELECT status, items FROM orders WHERE id=%s", (oid,))
            row = cur.fetchone()
            if row:
                old_status = row['status']
                old_items = row['items']
        if 'items' in data and isinstance(data['items'], list):
            data['items'] = json.dumps(data['items'])
        data = _filter_columns(data, ALLOWED_ORDER_COLUMNS)
        if data:
            sets = ', '.join(f"{k}=%s" for k in data)
            vals = list(data.values()) + [oid]
            cur.execute(f"UPDATE orders SET {sets} WHERE id=%s", vals)
        new_status = data.get('status', old_status)
        cancelled_statuses = ('cancelled', 'canceled')
        is_cancel = new_status in cancelled_statuses
        was_cancelled = old_status in cancelled_statuses

        if is_cancel and old_status and not was_cancelled:
            items_json = data.get('items') if 'items' in data else old_items
            if isinstance(items_json, str):
                items = json.loads(items_json)
            else:
                items = items_json
            for item in items:
                pid = item.get('product_id')
                qty = item.get('quantity') or item.get('qty') or 1
                color = item.get('color') or item.get('selectedColor') or ''
                size = item.get('size') or item.get('selectedSize') or ''
                if pid:
                    restore_order_stock(cur, pid, color, size, qty)

        elif was_cancelled and not is_cancel:
            items_json = data.get('items') if 'items' in data else old_items
            if isinstance(items_json, str):
                items = json.loads(items_json)
            else:
                items = items_json
            for idx, item in enumerate(items):
                pid = item.get('product_id')
                qty = item.get('quantity') or item.get('qty') or 1
                color = item.get('color') or item.get('selectedColor') or ''
                size = item.get('size') or item.get('selectedSize') or ''
                if pid:
                    stock_ok, err = deduct_order_stock(cur, pid, color, size, qty)
                    if not stock_ok:
                        for rollback_idx in range(idx):
                            prev = items[rollback_idx]
                            rpid = prev.get('product_id')
                            rqty = prev.get('quantity') or prev.get('qty') or 1
                            rcolor = prev.get('color') or prev.get('selectedColor') or ''
                            rsize = prev.get('size') or prev.get('selectedSize') or ''
                            if rpid:
                                restore_order_stock(cur, rpid, rcolor, rsize, rqty)
                        product_name = item.get('name', '')
                        msg = err or "Stock insuffisant"
                        if product_name:
                            msg = f"{msg} pour {product_name}"
                        db.commit()
                        return _json_response({'error': f"Impossible de réactiver la commande : {msg}"}, 409)

        if new_status and old_status and new_status != old_status:
            try:
                cur.execute("INSERT INTO status_history (order_id, status, note) VALUES (%s, %s, %s)",
                            (oid, new_status, ''))
            except Exception:
                pass
        db.commit()
        _signal_cache_invalidate()
        return {'message': 'Order updated'}
    finally:
        try:
            cur.close()
        except Exception:
            pass
        try:
            db.close()
        except Exception:
            pass


@router.put('/customers/{cid}')
def update_customer(cid: str, request: Request, data: dict = Body(...), session_token: str = Depends(require_admin_auth)):
    validate_admin_csrf(request)
    db = get_db()
    cur = db.cursor()
    try:
        if 'name' in data:
            data['name'] = data['name'].strip()
        data = _filter_columns(data, ALLOWED_CUSTOMER_COLUMNS)
        if not data:
            return {'message': 'Customer updated'}
        sets = ', '.join(f"{k}=%s" for k in data)
        vals = list(data.values()) + [cid]
        cur.execute(f"UPDATE customers SET {sets} WHERE id=%s", vals)
        db.commit()
        _signal_cache_invalidate()
        return {'message': 'Customer updated'}
    finally:
        try:
            cur.close()
        except Exception:
            pass
        try:
            db.close()
        except Exception:
            pass


@router.put('/inventory/{pid}')
def update_inventory(pid: str, request: Request, data: dict = Body(...), session_token: str = Depends(require_admin_auth)):
    validate_admin_csrf(request)
    db = get_db()
    cur = db.cursor()
    try:
        qty = data.get('quantity')
        if qty is not None:
            cur.execute("SELECT quantity FROM inventory WHERE product_id=%s", (pid,))
            before = cur.fetchone()
            before_qty = before['quantity'] if before else 0
            cur.execute("UPDATE inventory SET quantity=%s, updated_at=CURRENT_TIMESTAMP WHERE product_id=%s", (qty, pid))
            cur.execute("UPDATE products SET stock=%s WHERE id=%s", (qty, pid))
            log_stock_change(cur, pid, qty - before_qty, before_qty, data.get('reason', 'Manual update'))
            db.commit()
        _signal_cache_invalidate()
        return {'message': 'Inventory updated'}
    finally:
        try:
            cur.close()
        except Exception:
            pass
        try:
            db.close()
        except Exception:
            pass


# ---------------------------------------------------------------------------
# DELETE routes
# ---------------------------------------------------------------------------

@router.delete('/collections/{cid}')
def delete_collection(cid: str, request: Request, session_token: str = Depends(require_admin_auth)):
    validate_admin_csrf(request)
    db = get_db()
    cur = db.cursor()
    try:
        cur.execute("DELETE FROM collection_products WHERE collection_id=%s", (cid,))
        cur.execute("DELETE FROM collections WHERE id=%s", (cid,))
        db.commit()
        _signal_cache_invalidate()
        return {'message': 'Deleted'}
    finally:
        try:
            cur.close()
        except Exception:
            pass
        try:
            db.close()
        except Exception:
            pass


@router.delete('/products/delete-all')
def delete_all_products(request: Request, session_token: str = Depends(require_admin_auth)):
    validate_admin_csrf(request)
    db = get_db()
    cur = db.cursor()
    try:
        cur.execute("SELECT id, image, images FROM products")
        products = cur.fetchall()
        deleted_images = 0
        for p in products:
            pid = p['id']
            all_images = []
            if p['image']:
                all_images.append(p['image'])
            if p['images']:
                try:
                    imgs = json.loads(p['images']) if isinstance(p['images'], str) else p['images']
                    if isinstance(imgs, list):
                        all_images.extend(imgs)
                except Exception:
                    pass
            cur.execute("SELECT image_path FROM variant_images vi JOIN product_variants pv ON vi.variant_id = pv.id WHERE pv.product_id=%s", (pid,))
            for row in cur.fetchall():
                if row['image_path'] and row['image_path'] not in all_images:
                    all_images.append(row['image_path'])
            for img_path in all_images:
                if not img_path:
                    continue
                if storage.is_enabled() and (storage.is_cloudinary_url(img_path) or storage.is_supabase_url(img_path)):
                    sp = storage.path_from_url(img_path)
                    if sp:
                        storage.delete_file(sp)
                        deleted_images += 1
                elif not img_path.startswith('http'):
                    full_path = secure_path(PROJECT_ROOT, img_path)
                    if full_path and os.path.isfile(full_path):
                        os.remove(full_path)
                        deleted_images += 1
        cur.execute("DELETE FROM variant_images")
        cur.execute("DELETE FROM variant_sizes")
        cur.execute("DELETE FROM product_variants")
        cur.execute("DELETE FROM product_sizes")
        cur.execute("DELETE FROM product_colors")
        cur.execute("DELETE FROM collection_products")
        cur.execute("DELETE FROM inventory")
        cur.execute("DELETE FROM products")
        db.commit()
        _signal_cache_invalidate()
        return {'message': f'All products deleted. {deleted_images} local images removed.'}
    finally:
        try:
            cur.close()
        except Exception:
            pass
        try:
            db.close()
        except Exception:
            pass


@router.delete('/products/{pid}/images')
def delete_product_image(pid: str, request: Request, data: dict = Body({}), session_token: str = Depends(require_admin_auth)):
    validate_admin_csrf(request)
    db = get_db()
    cur = db.cursor()
    try:
        img_path = data.get('path', '')
        if img_path:
            if storage.is_enabled() and (storage.is_cloudinary_url(img_path) or storage.is_supabase_url(img_path)):
                sp = storage.path_from_url(img_path)
                if sp:
                    storage.delete_file(sp)
            elif not img_path.startswith('http'):
                full_path = secure_path(PROJECT_ROOT, img_path)
                if full_path and os.path.isfile(full_path):
                    os.remove(full_path)
        cur.execute("SELECT images FROM products WHERE id=%s", (pid,))
        row = cur.fetchone()
        images = json.loads(row['images']) if row and row['images'] else []
        if img_path in images:
            images.remove(img_path)
        primary = images[0] if images else ''
        cur.execute("UPDATE products SET images=%s, image=%s WHERE id=%s", (json.dumps(images), primary, pid))
        db.commit()
        _signal_cache_invalidate()
        return {'message': 'Image deleted', 'images': images}
    finally:
        try:
            cur.close()
        except Exception:
            pass
        try:
            db.close()
        except Exception:
            pass


@router.delete('/products/{pid}')
def delete_product(pid: str, request: Request, session_token: str = Depends(require_admin_auth)):
    validate_admin_csrf(request)
    db = get_db()
    cur = db.cursor()
    try:
        cur.execute("DELETE FROM inventory WHERE product_id=%s", (pid,))
        cur.execute("DELETE FROM products WHERE id=%s", (pid,))
        db.commit()
        _signal_cache_invalidate()
        return {'message': 'Deleted'}
    finally:
        try:
            cur.close()
        except Exception:
            pass
        try:
            db.close()
        except Exception:
            pass


@router.delete('/customers/{cid}')
def delete_customer(cid: str, request: Request, session_token: str = Depends(require_admin_auth)):
    validate_admin_csrf(request)
    db = get_db()
    cur = db.cursor()
    try:
        cur.execute("UPDATE orders SET customer_id=NULL WHERE customer_id=%s", (cid,))
        cur.execute("DELETE FROM customers WHERE id=%s", (cid,))
        db.commit()
        _signal_cache_invalidate()
        return {'message': 'Deleted'}
    finally:
        try:
            cur.close()
        except Exception:
            pass
        try:
            db.close()
        except Exception:
            pass


@router.delete('/categories/{cid}')
def delete_category(cid: str, request: Request, session_token: str = Depends(require_admin_auth)):
    validate_admin_csrf(request)
    db = get_db()
    cur = db.cursor()
    try:
        cur.execute("SELECT size_system FROM categories WHERE id=%s", (cid,))
        cat_row = cur.fetchone()
        if cat_row and cat_row['size_system'] == 'grouped_taille':
            return _json_response({'error': 'Cette catégorie est protégée (système de tailles groupées) et ne peut pas être supprimée.'}, 403)
        cur.execute("UPDATE products SET category_id=NULL WHERE category_id=%s", (cid,))
        cur.execute("DELETE FROM categories WHERE id=%s", (cid,))
        db.commit()
        _signal_cache_invalidate()
        return {'message': 'Deleted'}
    finally:
        try:
            cur.close()
        except Exception:
            pass
        try:
            db.close()
        except Exception:
            pass


@router.delete('/orders/{oid}')
def delete_order(oid: str, request: Request, session_token: str = Depends(require_admin_auth)):
    validate_admin_csrf(request)
    db = get_db()
    cur = db.cursor()
    try:
        cur.execute("SELECT status FROM orders WHERE id=%s", (oid,))
        row = cur.fetchone()
        if row is None:
            return _json_response({'error': 'Commande introuvable'}, 404)
        cur.execute("DELETE FROM orders WHERE id=%s", (oid,))
        db.commit()
        _signal_cache_invalidate()
        return {'message': 'Deleted'}
    finally:
        try:
            cur.close()
        except Exception:
            pass
        try:
            db.close()
        except Exception:
            pass
